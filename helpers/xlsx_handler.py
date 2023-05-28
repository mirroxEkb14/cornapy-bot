
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import config
import logger
import os

log = logger.get_logger(logger_name=__name__, file_name='logger/xlsx_handler.log')

class XlsxHandler:
	"""
	Contains methods for iteraction with the excel table
	It is used in admin mode when the admins want to see user's selection history
	"""

	def __init__(self):

		# the name of the excel file
		self.FILE_NAME = 'user_statistics.xlsx'
		self.wb = None

		self.SS_TITLE = 'Smart Search history'
		self.ss_ws = None

		self.SF_TITLE = 'Feedbacks history'
		self.sf_ws = None

	def create_excel(self):
		"""
		Used here in the 'write_ss_statistics()' function

		Creates a new excel file
		"""

		log.info("xlsx_call = create_excel()")


		def init_ss_sheet():
			self.ss_ws = self.wb.active
			self.ss_ws.title = self.SS_TITLE

			self.ss_ws.append(['tg_id', 'tg_name', 'date', 'mood', 'catalogue', 'genre'])
			self.style_first_row(self.ss_ws)

		def init_sf_sheet():
			self.sf_ws = self.wb.create_sheet(self.SF_TITLE)

			self.sf_ws.append(['tg_id', 'tg_name', 'movie', 'mood', 'verdict', 'comment'])
			self.style_first_row(self.sf_ws)


		self.wb = Workbook()
		init_ss_sheet()
		init_sf_sheet()
		self.wb.save(self.FILE_NAME)

	def style_first_row(self, ws):
		"""
		Used here in the 'init_ss_sheet()' and 'init_sf_sheet()' functions of the 'create_excel()' function
		Add some style to the names of the first six columns in the first row
		"""
		
		log.info("xlsx_call = style_first_row()")

		a1 = ws['A1']
		b1 = ws['B1']
		c1 = ws['C1']
		d1 = ws['D1']
		e1 = ws['E1']
		f1 = ws['F1']

		a1.font = Font(bold=True)
		b1.font = Font(bold=True)
		c1.font = Font(bold=True)
		d1.font = Font(bold=True)
		e1.font = Font(bold=True)
		f1.font = Font(bold=True)

	def write_data(self, users_search_history, users_feedback_history):
		"""
		During admin-mode in callbacks.py in the 'process_admin_excel()' function
		Returns 'FILE_NAME' so that in 'callbacks.py' we can send the file to the admin

		:param users_search_history = [
										(1, 1148695153, 'Daniyar, Daniyar', '29.10, 29.10', 'Боевое, Депрессия', 'Фильм, Фильм', 'Боевик, Драма'), 
										(2, 2060283357, 'Daniyar, Daniyar', '29.10, 30.10.22', 'Family, Депрессия', 'Movie, Фильм', 'Fiction, Драма')
									  ]

		:param users_feedback_history = [
											(1, 1148695153, 'Форрест гамп|Дорогой джон', 'Депрессия|Депрессия', 'Было не так уж и плохо |Смешанные чувства ', 'ggg|wwwwwwwwww'), 
											(2, 2060283357, 'Море соблазна', 'Депрессия', 'Смешанные чувства ', 'aaaaaaa')
										]
		"""
		log.info("xlsx_call = write_data()")

		if not os.path.exists(self.FILE_NAME):
			self.create_excel()
		else:
			os.remove(self.FILE_NAME)
			self.create_excel()

		self.write_ss_statistics(users_search_history)
		self.write_sf_statistics(users_feedback_history)

		return self.FILE_NAME


	"""
	-----
	The 'Smart Search history' sheet
	-----
	"""

	def write_ss_statistics(self, users_search_history):
		"""
		Writes ALL the data from 'user_search_history' db to an excel file

		:param data_list = [
								[
									1148695153, ['Daniyar', 'Daniyar'], ['03.11.22', '03.11.22'], ['Depression', 'Fighting'], ['Movie', 'Movie'], ['Drama', 'Action']
								], 
								[
									2060283357, ['Daniyar'], ['03.11.22'], ['Семья'], ['Сериал'], ['Триллер']
								]
							]
		"""

		log.info("xlsx_call = write_ss_statistics()")


		def get_data_list():
			"""Returns formatted 'user_search_history'"""
			
			formatted_list = []
			for inner_list_index in range(0, len(users_search_history)):
				inner_list = list(users_search_history[inner_list_index])
				inner_list.pop(0)
				formatted_list.append([inner_list[0]])
				
				for el in range(1, len(inner_list)):
					formatted_list[inner_list_index].append(inner_list[el].split(config.BotDB.USER_SEARCH_HISTORY_DELIMITER))

			return formatted_list


		data_list = get_data_list()

		self.wb = load_workbook(self.FILE_NAME)
		self.ss_ws = self.wb.active


		def write_to_ss_sheet():
			"""Writes each user to Excel, divides each of them with a blank row"""

			min_row_counter = 2
			max_row_counter = 2
			for inner_list_index in range(0, len(data_list)):

				min_row = min_row_counter
				min_col = 1

				max_row = len(data_list[inner_list_index][-1])
				max_col = len(data_list[inner_list_index])

				el_index = 1
				inside_el_index = 0

				for row in range(min_row, max_row + max_row_counter):
					for col in range(min_col, max_col + 1):
						char = get_column_letter(col)

						# write tg_id
						if row == min_row and col == 1:
							self.ss_ws[char + str(row)] = data_list[inner_list_index][0]

						# the left data
						elif char != 'A':
							self.ss_ws[char + str(row)] = data_list[inner_list_index][el_index][inside_el_index]
							el_index += 1

					el_index = 1
					inside_el_index += 1

				min_row_counter += (max_row + 1)
				max_row_counter = min_row_counter

			self.wb.save(self.FILE_NAME)


		write_to_ss_sheet()


	"""
	-----
	The 'Feedbacks history' sheet
	-----
	"""

	def write_sf_statistics(self, users_feedback_history):
		"""
		Writes ALL the data from 'old_feedbacks' db to an excel file

		:param data_list = [
								[
									1148695153, ['Daniyar', 'Daniyar'], ['Мой парень псих', '500 дней лета'], ['Депрессия', 'Депрессия'], ['Не то что я ожидал ', 'Смешанные чувства '], ['gg wp', 'gagaga']
								], 
								[
									2060283357, ['Daniyar'], ['Бойфренд из будущего'], ['Депрессия'], ['Not what I expected '], ['ezz']
								]
						   ]
		"""

		log.info("xlsx_call = write_sf_statistics()")

		def get_data_list():
			"""Returns formatted 'users_feedback_history'"""

			formatted_list = []
			for inner_list_index in range(0, len(users_feedback_history)):
				inner_list = list(users_feedback_history[inner_list_index])
				inner_list.pop(0)
				formatted_list.append([inner_list[0]])

				for el in range(1, len(inner_list)):
					formatted_list[inner_list_index].append(inner_list[el].split(config.BotDB.OLD_FEEDBACKS_DELIMITER))

			return formatted_list


		data_list = get_data_list()

		# self.wb = load_workbook(self.FILE_NAME)
		self.sf_ws = self.wb.get_sheet_by_name(self.SF_TITLE)


		def write_to_sf_sheet():
			"""Writes each user to Excel, divides each of them with a blank row"""

			min_row_counter = 2
			max_row_counter = 2
			for inner_list_index in range(0, len(data_list)):

				min_row = min_row_counter
				min_col = 1

				max_row = len(data_list[inner_list_index][-1])
				max_col = len(data_list[inner_list_index])

				el_index = 1
				inside_el_index = 0

				for row in range(min_row, max_row + max_row_counter):
					for col in range(min_col, max_col + 1):
						char = get_column_letter(col)

						# write tg_id
						if row == min_row and col == 1:
							self.sf_ws[char + str(row)] = data_list[inner_list_index][0]

						# the left data
						elif char != 'A':
							self.sf_ws[char + str(row)] = data_list[inner_list_index][el_index][inside_el_index]
							el_index += 1

					el_index = 1
					inside_el_index += 1

				min_row_counter += (max_row + 1)
				max_row_counter = min_row_counter

			self.wb.save(self.FILE_NAME)


		write_to_sf_sheet()
