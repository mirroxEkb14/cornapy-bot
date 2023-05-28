
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import gspread
import db

# BotDB = db.BotDB('cinema.db')

# sa = gspread.service_account()
# sh = sa.open('БД фильмов')

# wks = sh.worksheet('Sheet1')

# db_last_movie_id = BotDB.get_movies_last_id()
# google_sheet_movies = wks.get_all_records()

# print(google_sheet_movies[db_last_movie_id:])

input_list = [{'id': 1, 'name': 'name1'}, {'id': 2, 'name': 'name2'}]

print(list(input_list[0].values())[1:])


"""
-----------
"""

# users_feedback_history = [
#                             [
#                                 1148695153, ['Daniyar', 'Daniyar'], ['Мой парень псих', '500 дней лета'], ['Депрессия', 'Депрессия'], ['Не то что я ожидал ', 'Смешанные чувства '], ['gg wp', 'gagaga']
#                             ], 
#                             [
#                                 2060283357, ['Daniyar'], ['Бойфренд из будущего'], ['Депрессия'], ['Not what I expected '], ['ezz']
#                             ]
#                        ]

# wb = Workbook()
# ws = wb.active
# ws.title = 'Feedbacks history'

# ws.append(['tg_id', 'tg_name', 'movie', 'mood', 'verdict', 'comment'])
# wb.save('user_statistics.xlsx')

# def write_to_excel():
#     """Writes each user to Excel, divides each of them with a blank row"""

#     min_row_counter = 2
#     max_row_counter = 2
#     for inner_list_index in range(0, len(users_feedback_history)):

#         min_row = min_row_counter
#         min_col = 1

#         max_row = len(users_feedback_history[inner_list_index][-1])
#         max_col = len(users_feedback_history[inner_list_index])

#         el_index = 1
#         inside_el_index = 0

#         for row in range(min_row, max_row + max_row_counter):
#             for col in range(min_col, max_col + 1):
#                 char = get_column_letter(col)

#                 # write tg_id
#                 if row == min_row and col == 1:
#                     ws[char + str(row)] = users_feedback_history[inner_list_index][0]

#                 # the left data
#                 elif char != 'A':
#                     ws[char + str(row)] = users_feedback_history[inner_list_index][el_index][inside_el_index]
#                     el_index += 1

#             el_index = 1
#             inside_el_index += 1

#         min_row_counter += (max_row + 1)
#         max_row_counter = min_row_counter

#     wb.save('user_statistics.xlsx')

# write_to_excel()   


